from __future__ import annotations
from openpyxl import load_workbook
from typing import Dict


def netmask_to_bit_length(netmask: str) -> int:
    # 掩码到掩码长度
    """
    >>> netmask_to_bit_length('255.255.255.0')
    24
    >>>
    """
    # 分割字符串格式的子网掩码为四段列表
    # 计算二进制字符串中 '1' 的个数
    # 转换各段子网掩码为二进制, 计算十进制
    return sum([bin(int(i)).count('1') for i in netmask.split('.')])


def wildcard_mask_to_netmask(wildcard_mask: str) -> str:
    """
    反掩码到掩码
    >>> wildcard_mask_to_netmask('0.0.31.255')
    '255.255.224.0'
    >>>
    """
    return '.'.join([str(int(255) - int(i)) for i in wildcard_mask.split('.')])


def netmask_to_wildcard_mask(netmask: str) -> str:
    """
    掩码到反掩码
    >>> netmask_to_wildcard_mask('255.255.224.0')
    '0.0.31.255'
    """
    return wildcard_mask_to_netmask(netmask)


class ExcelDataGenerator:

    def __init__(self):
        self._data = {}  # type: Dict[str, Dict[str, str]]

    def load_generator(self, filename: str, sheetname: str = 'Sheet1', key=None):
        """
        加载excel文件, 这个是个迭代器
        @param filename: 文件名
        @param sheetname: sheet名
        @param key: 唯一字段名，如果为None，则按照第一行的字段名称自动生成
        """
        wb = load_workbook(filename)
        sheet = wb[sheetname]

        vars_list = []

        if key is None:
            key = sheet.cell(row=1, column=1).value

        # 取第一列中所有值，生成变量列表
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                vars_list.append(cell.value)

        # 对每一列进行遍历
        for col in range(3, sheet.max_column + 1):
            idx = 0
            ds = {}
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col, max_col=col):
                for cell in row:
                    ds[vars_list[idx]] = cell.value
                    idx += 1
            yield ds

            # 合并到所有
            self._data[ds[key]] = ds

    def load(self, filename: str, sheetname: str = 'Sheet1', key=None):
        """
        加载excel文件, 屏蔽掉了迭代器
        @param filename: 文件名
        @param sheetname: sheet名
        @param key: 唯一字段名，如果为None，则按照第一行的字段名称自动生成
        """
        for ds in self.load_generator(filename, sheetname, key):
            ...

    @property
    def data(self):
        """取从Excel中加载的数据"""
        return self._data.items()


def load_template(template_filename: str) -> str:
    """读取模板文件"""
    with open(template_filename, 'r+', encoding='utf-8', errors='ignore') as f:
        return f.read()


def write_template(template_str: str, output_file: str, data: dict):
    """
    将模板写入文件
    @param template_str: 模板字符串
    @param output_file: 输出文件
    @param data: 数据
    """
    with open(output_file, 'w+', encoding='utf-8') as f1:
        f1.write(template_str.format(**data))